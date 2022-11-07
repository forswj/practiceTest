import json
import jsonpath
import openpyxl
# from openpyxl import load_workbook
import pandas
from params_VAR.kanjia import EXCEL_PATH


# 读取excel内容，实现文件驱动自动化执行


def read_excel():
    excel = openpyxl.load_workbook(EXCEL_PATH)
    # excel = openpyxl.load_workbook('../data/api_cases_V4.xlsx')
    sheet = excel['Sheet1']
    # 创建装载Excel数据的变量
    tulpe_list = []
    # 逐行循环读取Excel数据
    for value in sheet.values:
        # 判断当前行的第一列的值，是否是数字编号
        if type(value[0]) is int:
            # 将元祖装进List
            tulpe_list.append(value)
    return tulpe_list


if __name__ == '__main__':
    print(read_excel())

'''
    1、通过输入的数字找到表中对应的行
    2、获取对应行第二列到最后一列的参数
    3、格式化成json格式并输出
'''
# 输入数字
# a = int(input('请输入应用对应第一列数字：'))

# 获取excel中的内容
# excel = openpyxl.load_workbook('../data/增购参数.xlsx')
# df = pandas.read_excel('../data/增购参数.xlsx')
#
# dataFile = []
# for a in range(df.values):
#     if a-1 > 0:
#         dataFile.append(df.iloc[a-1].values)
#
# print(dataFile)
